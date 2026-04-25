import pdfplumber
import openpyxl
from openpyxl import Workbook

class PDFProcessor:
def extract_tables_via_api(self, pdf_path, output_excel):
    """Использует PDF.co API для идеального распознавания таблиц"""
    import requests, json, time
    
    API_KEY = "t2468600@gmail.com_uDNhV6T9kk9ttSu5Enb22KGFYE9QGBnOGgTau84Vu63uSZKSJg61qHDYfHTyB4W3"  # Вставь сюда из pdf.co
    url = "https://api.pdf.co/v1/pdf/convert/to/excel"
    
    # Загружаем файл
    with open(pdf_path, 'rb') as f:
        files = {'file': f}
        headers = {'x-api-key': API_KEY}
        response = requests.post(url, files=files, headers=headers)
    
    result = response.json()
    
    if result.get('error'):
        return 0
    
    # Скачиваем готовый Excel
    excel_url = result['url']
    r = requests.get(excel_url)
    with open(output_excel, 'wb') as f:
        f.write(r.content)
    
    return 1  # API вернул Excel

 
 
 def extract_tables_to_excel(self, pdf_path, output_excel):
    from openpyxl import Workbook
    wb = Workbook()
    # Удаляем дефолтный пустой лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb.active)
    
    table_count = 0
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            for table_idx, table in enumerate(tables):
                if table and len(table) > 1:
                    table_count += 1
                    sheet_name = f"Page{page_num}_T{table_idx+1}"[:31]
                    ws = wb.create_sheet(title=sheet_name)
                    
                    for row_idx, row in enumerate(table):
                        for col_idx, cell in enumerate(row):
                            if cell:
                                ws.cell(row=row_idx+1, column=col_idx+1, value=str(cell))
    
    # Если нашли таблицы - сохраняем
    if table_count > 0:
        wb.save(output_excel)
    else:
        # Если таблиц нет - создаём пустой файл с одним листом
        wb2 = Workbook()
        wb2.active.title = "No tables found"
        wb2.save(output_excel)
    
    return table_count
    
    def find_explications(self, pdf_path):
        explications = []
        keywords = ['экспликация', 'помещение', 'комната', 'площадь', 'этаж']
        
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue
                    
                    table_text = ' '.join([' '.join([str(cell) for cell in row if cell]) for row in table]).lower()
                    
                    if any(kw in table_text for kw in keywords):
                        explications.append({
                            "page": page_num,
                            "rows": len(table)
                        })
        return explications
