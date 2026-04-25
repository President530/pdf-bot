import pdfplumber
import re
from openpyxl import Workbook

def extract_tables_to_excel_pro(pdf_path, output_excel):
    """PRO версия — как у ilovepdf"""
    
    import pdfplumber
    import re
    from openpyxl import Workbook
    
    wb = Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист
    
    sheet_count = 0
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            words = page.extract_words(keep_blank_chars=False)
            
            if not words:
                continue
            
            # Группируем по строкам (по Y координате)
            rows = {}
            threshold = 3  # Порог группировки в пунктах
            
            for w in words:
                y0 = round(w['y0'] / threshold) * threshold
                if y0 not in rows:
                    rows[y0] = []
                rows[y0].append(w)
            
            # Сортируем колонки в каждой строке по X
            table_rows = []
            for y in sorted(rows.keys()):
                row_words = sorted(rows[y], key=lambda x: x['x0'])
                row_text = [w['text'] for w in row_words]
                
                # Разбиваем склеенные числа
                expanded_row = []
                for cell in row_text:
                    if re.search(r'\d+\s+\d+', cell):
                        numbers = re.findall(r'\d+', cell)
                        expanded_row.extend(numbers)
                    else:
                        expanded_row.append(cell)
                
                table_rows.append(expanded_row)
            
            if len(table_rows) > 2:
                sheet_count += 1
                ws = wb.create_sheet(title=f"Страница_{page_num}")
                
                for row_idx, row in enumerate(table_rows):
                    for col_idx, cell in enumerate(row):
                        if cell:
                            ws.cell(row=row_idx+1, column=col_idx+1, value=cell)
    
    if sheet_count == 0:
        # Fallback к обычному методу
        return extract_tables_to_excel(pdf_path, output_excel)
    
    wb.save(output_excel)
    return sheet_count

def extract_tables_to_excel(pdf_path, output_excel):
    """Гибридный метод извлечения таблиц"""
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb.active)
    
    table_count = 0
    all_text = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            if not text:
                continue
            
            all_text.append(f"=== Страница {page_num} ===\n{text}")
            
            # Метод 1: стандартные таблицы pdfplumber
            tables = page.extract_tables()
            for table in tables:
                if table and len(table) > 1:
                    cleaned = []
                    for row in table:
                        if row and any(str(c).strip() for c in row if c):
                            cleaned.append([str(c).strip() if c else '' for c in row])
                    if len(cleaned) >= 2:
                        table_count += 1
                        ws = wb.create_sheet(title=f"Table_{table_count}")
                        for r_idx, row in enumerate(cleaned):
                            for c_idx, cell in enumerate(row):
                                if cell:
                                    ws.cell(row=r_idx+1, column=c_idx+1, value=cell)
            
            # Метод 2: строки с буквами и цифрами
            if table_count == 0:
                lines = text.split('\n')
                data_rows = []
                
                for line in lines:
                    has_cyrillic = re.search(r'[А-Яа-я]', line)
                    has_digits = re.search(r'\d', line)
                    
                    if has_cyrillic and has_digits:
                        cells = line.split()
                        if len(cells) >= 2:
                            data_rows.append(cells)
                
                if data_rows:
                    table_count = 1
                    ws = wb.create_sheet(title="Данные")
                    for r_idx, row in enumerate(data_rows):
                        for c_idx, cell in enumerate(row):
                            ws.cell(row=r_idx+1, column=c_idx+1, value=cell)
    
    if table_count == 0:
        ws = wb.create_sheet(title="Текст PDF")
        full_text = '\n'.join(all_text)
        for i, line in enumerate(full_text.split('\n')[:500]):
            if line.strip():
                ws.cell(row=i+1, column=1, value=line[:32767])
    
    wb.save(output_excel)
    return table_count if table_count > 0 else 0

def find_explications_smart(pdf_path):
    explications = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            
            for table in tables:
                if not table or len(table) < 2:
                    continue
                
                has_numbers = False
                has_names = False
                has_areas = False
                
                for row in table[:10]:
                    if not row:
                        continue
                    for cell in row:
                        if not cell:
                            continue
                        cell_str = str(cell).strip()
                        
                        if re.search(r'^\d+[\.\)]?\s*$|^\d+$', cell_str):
                            has_numbers = True
                        if re.search(r'[А-Я][а-я]+', cell_str) and len(cell_str) > 2:
                            has_names = True
                        if re.search(r'\d+[\.,]\d+\s*м?²?|\d+\s*м²', cell_str):
                            has_areas = True
                
                if has_numbers and has_names and has_areas:
                    formatted = []
                    for row in table:
                        if row and any(cell for cell in row):
                            formatted.append([str(cell).strip() if cell else '' for cell in row])
                    
                    explications.append({
                        'page': page_num,
                        'table': formatted,
                        'rows_count': len(formatted)
                    })
    
    return explications
