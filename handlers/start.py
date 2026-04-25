from handlers.pdf_utils import extract_tables_to_excel, find_explications_smart
import traceback
import os
import tempfile

print("🔧 DEBUG: handlers/start.py loading...")

# Хранилище PDF для каждого пользователя (временно)
user_pdfs = {}

def start_command(chat_id, send_message, get_keyboard):
    send_message(
        chat_id, 
        "🤖 *Здрасте, прывет от ВА*\n\n"
        "📌 *Что я умею:*\n"
        "• 📊 Извлекать таблицы из PDF в Excel\n"
        "• 📐 Находить экспликации помещений\n\n"
        "🚀 *Как работать:*\n"
        "1. Отправь мне PDF файл\n"
        "2. Нажми нужную кнопку в меню\n\n"
        "🆓 Бесплатно, без ограничений!",
        get_keyboard()
    )

def handle_document(chat_id, doc, send_message):
    """Сохраняет PDF и сообщает пользователю"""
    import requests
    import tempfile
    from app import URL, TOKEN
    
    send_message(chat_id, "📥 *Шаг 1/4: Скачиваю PDF...*")
    
    # Получаем файл
    send_message(chat_id, "📊 Получаю информацию о файле...")
    file_info = requests.get(URL + f"/getFile?file_id={doc['file_id']}").json()
    
    if not file_info.get('ok'):
        send_message(chat_id, "❌ Ошибка получения файла")
        return
    
    file_path = file_info['result']['file_path']
    file_url = f"https://api.telegram.org/file/bot{TOKEN}/{file_path}"
    
    # Скачиваем
    send_message(chat_id, "📥 Скачиваю содержимое...")
    r = requests.get(file_url)
    temp_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    temp_pdf.write(r.content)
    temp_pdf.close()
    
    # Сохраняем путь
    user_pdfs[chat_id] = temp_pdf.name
    file_size = os.path.getsize(temp_pdf.name) / 1024  # в КБ
    
    send_message(
        chat_id, 
        f"✅ *PDF принят!*\n\n"
        f"📄 Имя: {doc.get('file_name', 'без имени')}\n"
        f"📦 Размер: {file_size:.1f} КБ\n"
        f"🔗 Путь: {temp_pdf.name}\n\n"
        f"📌 Теперь выбери действие в меню:",
        get_keyboard()
    )

def handle_text(chat_id, text, send_message, send_document):
    import os
    import tempfile
    import time
    
    # Отправляем начальное сообщение
    send_message(chat_id, f"🔍 *Обработка команды:* {text}")
    
    # Проверка наличия PDF
    if chat_id not in user_pdfs:
        send_message(chat_id, "❌ *Сначала отправь PDF файл!*")
        return
    
    pdf_path = user_pdfs[chat_id]
    send_message(chat_id, f"✅ PDF найден: {os.path.basename(pdf_path)}")
    
    # Стандартный режим
    if text == '📊 Таблицы в Excel' or text == '/tables':
        send_message(chat_id, "⏳ *Запускаю стандартный режим...*")
        
        output_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
        send_message(chat_id, "📊 Извлекаю таблицы (это может занять 5-10 секунд)...")
        
        start_time = time.time()
        count = extract_tables_to_excel(pdf_path, output_excel)
        elapsed = time.time() - start_time
        
        send_message(chat_id, f"⏱ Время обработки: {elapsed:.1f} сек")
        
        if count == 0:
            send_message(chat_id, "❌ *Таблицы не найдены* в этом PDF.")
        else:
            send_message(chat_id, f"✅ Найдено {count} таблиц, создаю Excel файл...")
            send_document(chat_id, output_excel, f"tables_{count}.xlsx")
            os.unlink(output_excel)
            send_message(chat_id, "✅ Готово!")
    
    # Режим экспликаций
    elif text == '📐 Экспликации' or text == '/explication':
        send_message(chat_id, "🔍 *Запускаю поиск экспликаций...*")
        
        start_time = time.time()
        result = find_explications_smart(pdf_path)
        elapsed = time.time() - start_time
        
        send_message(chat_id, f"⏱ Время обработки: {elapsed:.1f} сек")
        
        if not result:
            send_message(chat_id, "❌ *Экспликации не найдены* в этом PDF.\n\n"
                                  "💡 Совет: убедись что в файле есть таблица с номерами, названиями и площадями комнат.")
        else:
            msg = f"🔍 *Найдено {len(result)} таблиц с экспликациями:*\n\n"
            for r in result:
                msg += f"📄 *Страница {r['page']}* — {r['rows_count']} строк\n"
                # Показываем пример первой строки
                if r['table'] and len(r['table']) > 0:
                    first_row = r['table'][0]
                    msg += f"   Пример: {' | '.join([str(c)[:15] for c in first_row if c])}\n"
                msg += "\n"
            
            if len(msg) > 4000:
                msg = msg[:4000] + "\n\n...(обрезано)"
            
            send_message(chat_id, msg)
            send_message(chat_id, "💡 Чтобы получить полный Excel с экспликациями, используй режим 'Таблицы в Excel'")
    
    # PRO режим (подробная отладка)
    elif text == '🚀 Excel (PRO)':
        send_message(chat_id, "🚀 *ЗАПУСК PRO РЕЖИМА С ОТЛАДКОЙ*\n")
        
        # Шаг 1: Проверка файла
        send_message(chat_id, "📋 **ШАГ 1/6: Проверка файла**")
        if not os.path.exists(pdf_path):
            send_message(chat_id, f"❌ Файл НЕ НАЙДЕН: {pdf_path}")
            return
        file_size = os.path.getsize(pdf_path)
        send_message(chat_id, f"✅ Файл найден\n📦 Размер: {file_size} байт ({file_size/1024:.1f} КБ)")
        
        # Шаг 2: Проверка функции
        send_message(chat_id, "\n📋 **ШАГ 2/6: Проверка PRO-функции**")
        if 'extract_tables_to_excel_pro' not in globals():
            send_message(chat_id, "❌ PRO-функция не определена в глобальной области!")
            send_message(chat_id, "🔍 Доступные функции: " + ", ".join([k for k in globals().keys() if not k.startswith('_')])[:200])
            return
        send_message(chat_id, "✅ PRO-функция найдена")
        
        # Шаг 3: Подготовка выходного файла
        send_message(chat_id, "\n📋 **ШАГ 3/6: Создание временного файла**")
        output_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
        send_message(chat_id, f"✅ Временный файл: {os.path.basename(output_excel)}")
        
        # Шаг 4: Запуск обработки
        send_message(chat_id, "\n📋 **ШАГ 4/6: Запуск обработки PDF**")
        send_message(chat_id, "⏳ Это может занять 10-30 секунд, пожалуйста, подождите...")
        
        start_time = time.time()
        try:
            count = extract_tables_to_excel_pro(pdf_path, output_excel)
            elapsed = time.time() - start_time
            
            send_message(chat_id, f"\n📋 **ШАГ 5/6: Результат обработки**")
            send_message(chat_id, f"⏱ Время: {elapsed:.1f} секунд")
            send_message(chat_id, f"📊 Найдено таблиц: {count}")
            
            if count == 0:
                send_message(chat_id, "❌ Таблицы не найдены. Попробуйте стандартный режим.")
                os.unlink(output_excel)
            else:
                send_message(chat_id, f"\n📋 **ШАГ 6/6: Отправка файла**")
                send_message(chat_id, f"✅ Создаю Excel с {count} таблиц(ами)...")
                
                # Проверяем размер выходного файла
                if os.path.exists(output_excel):
                    out_size = os.path.getsize(output_excel) / 1024
                    send_message(chat_id, f"📦 Размер Excel: {out_size:.1f} КБ")
                
                send_document(chat_id, output_excel, f"pro_tables_{count}_{int(elapsed)}sec.xlsx")
                os.unlink(output_excel)
                send_message(chat_id, "✅ *ГОТОВО!* PRO-обработка успешно завершена.")
                
        except Exception as e:
            elapsed = time.time() - start_time
            send_message(chat_id, f"\n❌ **ОШИБКА на шаге 4** (через {elapsed:.1f} сек)")
            send_message(chat_id, f"Тип ошибки: {type(e).__name__}")
            send_message(chat_id, f"Сообщение: {str(e)[:300]}")
            
            # Отправляем traceback в лог (но не в Telegram, чтобы не заспамить)
            error_details = traceback.format_exc()
            print(f"PRO ERROR DETAILS:\n{error_details}")
            send_message(chat_id, "\n🔍 Подробности ошибки отправлены в лог сервера.")
            send_message(chat_id, "💡 Попробуйте стандартный режим 'Таблицы в Excel'")

def get_keyboard():
    from keyboards.menu import main_menu_keyboard
    return main_menu_keyboard()

# ========== PRO ФУНКЦИЯ (улучшенная) ==========
import pdfplumber
import re
from openpyxl import Workbook

def extract_tables_to_excel_pro(pdf_path, output_excel):
    """Экономичная PRO версия (без extract_words на больших PDF)"""
    wb = Workbook()
    wb.remove(wb.active)
    sheet_count = 0
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Ограничиваем количество страниц для бесплатного тарифа
            max_pages = min(len(pdf.pages), 15)
            
            for page_num in range(max_pages):
                page = pdf.pages[page_num]
                
                # Метод 1: Пробуем extract_tables (самый быстрый и экономный)
                tables = page.extract_tables()
                found_tables = False
                
                for table in tables:
                    if table and len(table) >= 2:
                        # Очищаем и обрабатываем таблицу
                        processed_rows = []
                        for row_idx, row in enumerate(table[:50]):  # Ограничиваем строки
                            if not row or not any(cell for cell in row):
                                continue
                            
                            new_row = []
                            for cell in row:
                                if cell and isinstance(cell, str):
                                    cell_clean = cell.strip()
                                    # Разбиваем только явные склеенные числа
                                    if re.search(r'\d+\s+\d+', cell_clean):
                                        parts = cell_clean.split()
                                        # Разбиваем только если части выглядят как числа
                                        if all(p.isdigit() for p in parts):
                                            new_row.extend(parts)
                                            continue
                                    new_row.append(cell_clean)
                                elif cell:
                                    new_row.append(str(cell))
                                else:
                                    new_row.append('')
                            
                            if new_row and any(new_row):
                                processed_rows.append(new_row)
                        
                        if len(processed_rows) >= 2:
                            found_tables = True
                            sheet_count += 1
                            ws = wb.create_sheet(title=f"Page_{page_num+1}_{sheet_count}")
                            
                            # Записываем только первые 100 строк для экономии памяти
                            for i, row in enumerate(processed_rows[:100]):
                                for j, val in enumerate(row[:20]):  # Ограничиваем колонки
                                    if val:
                                        ws.cell(row=i+1, column=j+1, value=val)
                
                # Метод 2: Если таблиц нет, пробуем просто текст (но экономно)
                if not found_tables:
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        data_rows = []
                        for line in lines[:100]:  # Ограничиваем строки
                            if line.strip() and any(c.isdigit() for c in line):
                                cells = line.split()
                                if len(cells) >= 3:  # Минимум 3 колонки
                                    data_rows.append(cells)
                        
                        if len(data_rows) >= 3:
                            sheet_count += 1
                            ws = wb.create_sheet(title=f"Page_{page_num+1}_text")
                            for i, row in enumerate(data_rows[:50]):
                                for j, val in enumerate(row[:15]):
                                    if val:
                                        ws.cell(row=i+1, column=j+1, value=val)
                
                # Принудительно очищаем память после каждой страницы
                del page
                
            # Очищаем память от PDF
            pdf.close()
    
    except Exception as e:
        print(f"PRO ERROR: {e}")
        import traceback
        traceback.print_exc()
        return extract_tables_to_excel(pdf_path, output_excel)
    
    if sheet_count == 0:
        print("No tables found, falling back to standard method")
        return extract_tables_to_excel(pdf_path, output_excel)
    
    try:
        wb.save(output_excel)
    except Exception as e:
        print(f"Save error: {e}")
        return 0
    
    return sheet_count

print("✅ handlers/start.py успешно загружен")
print(f"✅ extract_tables_to_excel_pro определена: {extract_tables_to_excel_pro is not None}")
